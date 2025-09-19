import os
import io
import base64
import logging
import json
from typing import List, Dict

from azure.core.credentials import AzureKeyCredential
from azure.search.documents import SearchClient
from azure.search.documents.indexes import SearchIndexClient
from azure.search.documents.indexes.models import (
    SearchIndex,
    SimpleField,
    SearchFieldDataType,
    SearchableField,
    CorsOptions,
    SemanticConfiguration,
    SemanticPrioritizedFields,
    SemanticField,
    SemanticSearch,
    # Added for vector search
    VectorSearch,
    HnswAlgorithmConfiguration,
    VectorSearchProfile,
    SearchField,
)
from azure.storage.blob import BlobServiceClient
from azure.identity import DefaultAzureCredential

# -----------------------------
# Environment configuration
# -----------------------------
SEARCH_ENDPOINT = os.environ.get("AZURE_SEARCH_ENDPOINT", "")
SEARCH_INDEX = os.environ.get("AZURE_SEARCH_INDEX", "")  # fallback / legacy
SEARCH_KEY = os.environ.get("AZURE_SEARCH_ADMIN_KEY", "")
SEM_CONFIG_NAME = os.environ.get("AZURE_SEARCH_SEMANTIC_CONFIG")

# Storage: prefer connection string, then account URL (optionally with SAS), else MSI
AZ_STORAGE_CONN_STR = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
AZ_STORAGE_ACCOUNT_URL = os.getenv("AZURE_STORAGE_ACCOUNT_URL")
# Add storage account name for constructing URL if needed
AZ_STORAGE_ACCOUNT_NAME = os.getenv("AZURE_STORAGE_ACCOUNT_NAME")

# Azure OpenAI (embeddings) configuration
AOAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT", "")
AOAI_API_KEY = os.getenv("AZURE_OPENAI_API_KEY", "")
AOAI_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION", "2024-10-01-preview")
AOAI_EMBED_DEPLOYMENT = os.getenv("AZURE_OPENAI_EMBEDDING_DEPLOYMENT", "")
# default to 3072 for text-embedding-3-large
AOAI_EMBED_DIM = int(os.getenv("AZURE_OPENAI_EMBED_DIM", "3072"))


# -----------------------------
# Helpers matching existing repo
# -----------------------------

def _blob_client() -> BlobServiceClient:
    """
    Create BlobServiceClient with the following priority:
    1. Connection string (if provided for backward compatibility)
    2. Account URL with Managed Identity (recommended)
    3. Account name to construct URL with Managed Identity
    """
    # Option 1: Use connection string if provided (backward compatibility)
    if AZ_STORAGE_CONN_STR:
        logging.info("Using storage connection string for authentication")
        return BlobServiceClient.from_connection_string(AZ_STORAGE_CONN_STR)
    
    # Option 2: Use account URL with Managed Identity or SAS
    if AZ_STORAGE_ACCOUNT_URL:
        # If SAS token is in URL (contains '?'), use it directly
        if "?" in AZ_STORAGE_ACCOUNT_URL:
            logging.info("Using storage account URL with SAS token")
            return BlobServiceClient(account_url=AZ_STORAGE_ACCOUNT_URL)
        # Otherwise use Managed Identity
        logging.info("Using storage account URL with Managed Identity")
        return BlobServiceClient(
            account_url=AZ_STORAGE_ACCOUNT_URL, 
            credential=DefaultAzureCredential()
        )
    
    # Option 3: Construct URL from account name and use Managed Identity
    if AZ_STORAGE_ACCOUNT_NAME:
        account_url = f"https://{AZ_STORAGE_ACCOUNT_NAME}.blob.core.windows.net"
        logging.info(f"Using constructed storage URL with Managed Identity: {account_url}")
        return BlobServiceClient(
            account_url=account_url,
            credential=DefaultAzureCredential()
        )
    
    # If none are set, raise an error
    raise RuntimeError(
        "Storage configuration missing. Set one of: "
        "AZURE_STORAGE_CONNECTION_STRING, AZURE_STORAGE_ACCOUNT_URL, or AZURE_STORAGE_ACCOUNT_NAME"
    )


def _get_aoai_client():
    """Create Azure OpenAI client using OpenAI SDK (AzureOpenAI). Returns None if not configured."""
    if not AOAI_ENDPOINT or not AOAI_API_KEY or not AOAI_EMBED_DEPLOYMENT:
        logging.warning("Azure OpenAI not fully configured; proceeding without embeddings")
        return None
    try:
        from openai import AzureOpenAI  # provided by openai>=1.x
        return AzureOpenAI(api_key=AOAI_API_KEY, api_version=AOAI_API_VERSION, azure_endpoint=AOAI_ENDPOINT)
    except Exception as e:  # pragma: no cover
        logging.exception("Failed to create AzureOpenAI client: %s", e)
        return None


def _embed_texts(texts: List[str]) -> List[List[float] | None]:
    """Return embeddings for texts via AzureOpenAI. If not available, returns [None,...]."""
    client = _get_aoai_client()
    if not client:
        return [None] * len(texts)

    vectors: List[List[float] | None] = []
    batch_size = 16
    for i in range(0, len(texts), batch_size):
        batch = texts[i : i + batch_size]
        try:
            resp = client.embeddings.create(model=AOAI_EMBED_DEPLOYMENT, input=batch)
            vecs = [d.embedding for d in resp.data]
            for v in vecs:
                if isinstance(v, list) and len(v) != AOAI_EMBED_DIM:
                    logging.warning("Embedding dimension %s != expected %s", len(v), AOAI_EMBED_DIM)
            vectors.extend(vecs)
        except Exception as e:  # pragma: no cover
            logging.exception("Embedding request failed: %s", e)
            vectors.extend([None] * len(batch))
    return vectors


def _sanitize_index_name(app_id: str) -> str:
    """Convert arbitrary appId to a valid Azure AI Search index name.

    Rules: lowercase, alphanumerics or dashes; must start/end with alphanumeric; length 2-128.
    """
    import re, hashlib
    base = app_id.lower().strip()
    base = re.sub(r"[^a-z0-9-]", "-", base)            # invalid chars -> dash
    base = re.sub(r"-+", "-", base)                     # collapse dashes
    base = base.strip("-")                                 # trim
    if not base or not base[0].isalnum():
        base = f"app-{hashlib.sha1(app_id.encode()).hexdigest()[:8]}"
    if len(base) < 2:
        base = f"{base}ix"
    if len(base) > 128:
        base = base[:128]
    return base


def create_or_update_index(app_id: str, force_recreate: bool = True) -> str:
    """Create (or recreate) a search index whose name is derived from app_id.

    If force_recreate=True and the index exists, delete it first so schema changes are applied.
    Returns the index name actually used.
    """
    if not SEARCH_ENDPOINT:
        raise RuntimeError("AZURE_SEARCH_ENDPOINT must be set")
    if not SEARCH_KEY:
        logging.warning("AZURE_SEARCH_ADMIN_KEY not set; index creation may fail without permissions")

    index_name = _sanitize_index_name(app_id)
    idx_client = SearchIndexClient(SEARCH_ENDPOINT, AzureKeyCredential(SEARCH_KEY))

    fields = [
        SimpleField(name="id", type=SearchFieldDataType.String, key=True),
        SimpleField(name="appId", type=SearchFieldDataType.String, filterable=True, sortable=True, facetable=True),
        SearchableField(name="title", type=SearchFieldDataType.String, analyzer_name="en.lucene"),
        SearchableField(name="content", type=SearchFieldDataType.String, analyzer_name="en.lucene"),
        SimpleField(name="source", type=SearchFieldDataType.String, filterable=True),
        SimpleField(name="path", type=SearchFieldDataType.String, filterable=True),
        SimpleField(name="chunkId", type=SearchFieldDataType.String, filterable=True),
    ]

    # Add vector field only if embeddings configured (dimension > 0) else skip to avoid errors.
    if AOAI_EMBED_DIM > 0:
        fields.append(
            SearchField(
                name="contentVector",
                type=SearchFieldDataType.Collection(SearchFieldDataType.Single),
                searchable=True,
                vector_search_dimensions=AOAI_EMBED_DIM,
                vector_search_profile_name="v1",
            )
        )

    index = SearchIndex(
        name=index_name,
        fields=fields,
        cors_options=CorsOptions(allowed_origins=["*"], max_age_in_seconds=60),
    )

    # Vector search config if vector field present
    if any(f.name == "contentVector" for f in fields):
        index.vector_search = VectorSearch(
            algorithms=[HnswAlgorithmConfiguration(name="hnsw")],
            profiles=[VectorSearchProfile(name="v1", algorithm_configuration_name="hnsw")],
        )

    if SEM_CONFIG_NAME:
        index.semantic_search = SemanticSearch(
            configurations=[
                SemanticConfiguration(
                    name=SEM_CONFIG_NAME,
                    prioritized_fields=SemanticPrioritizedFields(
                        title_field=SemanticField(field_name="title"),
                        content_fields=[SemanticField(field_name="content")],
                    ),
                )
            ]
        )

    # Recreate if required
    try:
        existing = idx_client.get_index(index_name)
        if existing and force_recreate:
            logging.info("Deleting existing index '%s' to recreate", index_name)
            idx_client.delete_index(index_name)
    except Exception:
        # Not found -> safe to create
        pass

    # Create (again) - if it still exists and force_recreate False, this will raise; catch & ignore
    try:
        idx_client.create_index(index)
        logging.info("Index '%s' created", index_name)
    except Exception as e:
        logging.debug("Index create may have failed or already exists: %s", e)
    return index_name


def _download_text(container: str, blob_name: str) -> str:
    bc = _blob_client().get_blob_client(container=container, blob=blob_name)
    data = bc.download_blob().readall()
    name = blob_name.lower()
    
    # Existing file type handlers...
    if name.endswith((".txt", ".md")):
        return data.decode("utf-8", errors="ignore")
    if name.endswith(".docx"):
        from docx import Document
        f = io.BytesIO(data)
        doc = Document(f)
        return "\n".join([p.text for p in doc.paragraphs])
    if name.endswith(".pdf"):
        from pypdf import PdfReader
        f = io.BytesIO(data)
        reader = PdfReader(f)
        return "\n".join([(page.extract_text() or "") for page in reader.pages])
    
    # Enhanced Excel handling for dependency files
    if name.endswith(".xlsx"):
        # Check if this is a dependency file based on name or content
        is_dependency_file = any(keyword in name for keyword in ['dependency', 'dependencies', 'connection', 'network'])
        
        if not is_dependency_file:
            # Quick check for dependency-related headers
            from openpyxl import load_workbook
            f = io.BytesIO(data)
            wb = load_workbook(f, read_only=True, data_only=True)
            for ws in wb.worksheets:
                for row in ws.iter_rows(max_row=5, values_only=True):
                    row_text = ' '.join(str(cell).lower() for cell in row if cell)
                    if 'source server' in row_text or 'destination server' in row_text:
                        is_dependency_file = True
                        break
                if is_dependency_file:
                    break
        
        if is_dependency_file:
            # Parse as dependency data
            records = _parse_dependency_excel(data, container, blob_name)
            if records:
                chunks = _create_dependency_chunks(records)
                return "\n\n---DEPENDENCY_CHUNK---\n\n".join(chunks)
        
        # Default Excel handling for non-dependency files
        from openpyxl import load_workbook
        f = io.BytesIO(data)
        wb = load_workbook(f, read_only=True, data_only=True)
        lines: List[str] = []
        for ws in wb.worksheets:
            lines.append(f"# Sheet: {ws.title}")
            for row in ws.iter_rows(values_only=True):
                values = [(str(c) if c is not None else "") for c in row]
                line = " \t ".join(v.strip() for v in values if v is not None)
                if line.strip():
                    lines.append(line)
        return "\n".join(lines)
    
    # Fallback: treat as text
    return data.decode("utf-8", errors="ignore")


def _chunk(text: str, size: int = 1200, overlap: int = 200) -> List[str]:
    words = text.split()
    chunks: List[str] = []
    i = 0
    while i < len(words):
        chunk_words = words[i : i + size]
        chunks.append(" ".join(chunk_words))
        i += max(1, size - overlap)
    return chunks


def _safe_doc_id(app_id: str, path: str, ci: int) -> str:
    raw = f"{app_id}|{path}|{ci}".encode("utf-8")
    return base64.urlsafe_b64encode(raw).decode("ascii")


def _parse_dependency_excel(data: bytes, container: str, blob_name: str) -> List[Dict]:
    """Parse dependency Excel file into structured records."""
    from openpyxl import load_workbook
    import hashlib
    
    f = io.BytesIO(data)
    wb = load_workbook(f, read_only=True, data_only=True)
    
    dependency_records = []
    
    for ws in wb.worksheets:
        headers = []
        header_row_found = False
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            # Look for header row containing key columns
            if not header_row_found:
                row_values = [str(cell).strip().lower() if cell else "" for cell in row]
                if any(col in ' '.join(row_values) for col in ['source server', 'destination server', 'source ip']):
                    headers = [str(cell).strip() if cell else "" for cell in row]
                    header_row_found = True
                    continue
            
            if header_row_found and any(cell for cell in row):
                # Create structured record from row data
                record = {}
                for idx, value in enumerate(row):
                    if idx < len(headers) and headers[idx] and value:
                        record[headers[idx]] = str(value).strip()
                
                if record:  # Only add non-empty records
                    # Create a unique ID for this dependency record
                    record_str = json.dumps(record, sort_keys=True)
                    record_hash = hashlib.md5(record_str.encode()).hexdigest()[:8]
                    record['_id'] = f"{container}_{blob_name}_{record_hash}"
                    record['_source_file'] = f"{container}/{blob_name}"
                    dependency_records.append(record)
    
    return dependency_records

def _create_dependency_chunks(records: List[Dict]) -> List[str]:
    """Create searchable text chunks from dependency records."""
    chunks = []
    
    for record in records:
        # Create a semantic representation of the dependency
        parts = []
        
        # Build a natural language description
        if record.get('Time slot'):
            parts.append(f"Time: {record['Time slot']}")
        
        # Source information
        src_desc = []
        if record.get('Source server name'):
            src_desc.append(f"Server: {record['Source server name']}")
        if record.get('Source IP'):
            src_desc.append(f"IP: {record['Source IP']}")
        if record.get('Source application'):
            src_desc.append(f"Application: {record['Source application']}")
        if record.get('Source process'):
            src_desc.append(f"Process: {record['Source process']}")
        
        if src_desc:
            parts.append(f"SOURCE - {', '.join(src_desc)}")
        
        # Destination information
        dst_desc = []
        if record.get('Destination server name'):
            dst_desc.append(f"Server: {record['Destination server name']}")
        if record.get('Destination IP'):
            dst_desc.append(f"IP: {record['Destination IP']}")
        if record.get('Destination application'):
            dst_desc.append(f"Application: {record['Destination application']}")
        if record.get('Destination process'):
            dst_desc.append(f"Process: {record['Destination process']}")
        if record.get('Destination port'):
            dst_desc.append(f"Port: {record['Destination port']}")
        
        if dst_desc:
            parts.append(f"DESTINATION - {', '.join(dst_desc)}")
        
        # Create the chunk text
        chunk_text = " | ".join(parts)
        
        # Add structured JSON at the end for precise extraction
        chunk_text += f"\n[DEPENDENCY_DATA]{json.dumps(record, separators=(',', ':'))}[/DEPENDENCY_DATA]"
        
        chunks.append(chunk_text)
    
    # Group chunks if they're too small
    grouped_chunks = []
    current_group = []
    current_size = 0
    max_chunk_size = 600  # Smaller chunks for dependency data
    
    for chunk in chunks:
        chunk_size = len(chunk.split())
        if current_size + chunk_size > max_chunk_size and current_group:
            grouped_chunks.append("\n\n".join(current_group))
            current_group = [chunk]
            current_size = chunk_size
        else:
            current_group.append(chunk)
            current_size += chunk_size
    
    if current_group:
        grouped_chunks.append("\n\n".join(current_group))
    
    return grouped_chunks

# -----------------------------
# Public functions called by HTTP
# -----------------------------

# Batch controls: Azure Search REST cap is 1000 docs or ~16MB per request.
MAX_DOCS_PER_BATCH = int(os.getenv("SEARCH_MAX_DOCS_PER_BATCH", "500"))
# Keep some headroom below 16MB to account for HTTP and JSON overhead
MAX_BYTES_PER_BATCH = int(os.getenv("SEARCH_MAX_BYTES_PER_BATCH", str(12 * 1024 * 1024)))

def _estimate_doc_bytes(doc: Dict) -> int:
    # Estimate serialized size of a single doc
    # ensure_ascii=False keeps unicode; separators minimize overhead
    s = json.dumps(doc, ensure_ascii=False, separators=(",", ":"))
    return len(s.encode("utf-8"))

def _flush_batch(search: SearchClient, batch: List[Dict]) -> Dict[str, int]:
    if not batch:
        return {"uploaded": 0, "failed": 0}
    resp = search.upload_documents(documents=batch)
    uploaded = sum(1 for r in resp if r.succeeded)
    failed = len(resp) - uploaded
    return {"uploaded": uploaded, "failed": failed}

def index_blob(app_id: str, container: str, blob_name: str) -> Dict:
    # Ensure (re)created index for this app
    index_name = create_or_update_index(app_id)

    search = SearchClient(SEARCH_ENDPOINT, index_name, AzureKeyCredential(SEARCH_KEY))

    text = _download_text(container, blob_name)
    if not text or not text.strip():
        return {"blobName": blob_name, "chunks": 0, "uploaded": 0, "failed": 0}

    chunks = _chunk(text)
    vectors = _embed_texts(chunks)

    batch: List[Dict] = []
    batch_bytes = 0
    uploaded = 0
    failed = 0

    for ci, ch in enumerate(chunks):
        path = f"{container}/{blob_name}"
        doc = {
            "id": _safe_doc_id(app_id, path, ci),
            "appId": app_id,
            "title": os.path.basename(blob_name),
            "content": ch,
            "source": "blob",
            "path": path,
            "chunkId": f"{ci}",
        }
        # attach embedding if available
        vec = vectors[ci] if vectors and ci < len(vectors) else None
        if isinstance(vec, list):
            doc["contentVector"] = [float(x) for x in vec]
        size = _estimate_doc_bytes(doc)
        if (len(batch) + 1 > MAX_DOCS_PER_BATCH) or (batch_bytes + size > MAX_BYTES_PER_BATCH):
            res = _flush_batch(search, batch)
            uploaded += res["uploaded"]
            failed += res["failed"]
            batch = []
            batch_bytes = 0
        batch.append(doc)
        batch_bytes += size

    if batch:
        res = _flush_batch(search, batch)
        uploaded += res["uploaded"]
        failed += res["failed"]

    return {"blobName": blob_name, "chunks": len(chunks), "uploaded": uploaded, "failed": failed}


def index_container(app_id: str, container: str) -> Dict:
    # Ensure (re)created index for this app
    index_name = create_or_update_index(app_id)

    search = SearchClient(SEARCH_ENDPOINT, index_name, AzureKeyCredential(SEARCH_KEY))
    bs = _blob_client().get_container_client(container)

    total_uploaded = 0
    total_failed = 0
    total_chunks = 0
    total_blobs = 0

    batch: List[Dict] = []
    batch_bytes = 0

    for blob in bs.list_blobs():
        name = blob.name
        try:
            text = _download_text(container, name)
            if not text or not text.strip():
                continue
            chunks = _chunk(text)
            total_chunks += len(chunks)
            total_blobs += 1

            vectors = _embed_texts(chunks)

            for ci, ch in enumerate(chunks):
                path = f"{container}/{name}"
                doc = {
                    "id": _safe_doc_id(app_id, path, ci),
                    "appId": app_id,
                    "title": os.path.basename(name),
                    "content": ch,
                    "source": "blob",
                    "path": path,
                    "chunkId": f"{ci}",
                }
                vec = vectors[ci] if vectors and ci < len(vectors) else None
                if isinstance(vec, list):
                    doc["contentVector"] = [float(x) for x in vec]
                size = _estimate_doc_bytes(doc)
                if (len(batch) + 1 > MAX_DOCS_PER_BATCH) or (batch_bytes + size > MAX_BYTES_PER_BATCH):
                    res = _flush_batch(search, batch)
                    total_uploaded += res["uploaded"]
                    total_failed += res["failed"]
                    batch = []
                    batch_bytes = 0
                batch.append(doc)
                batch_bytes += size
        except Exception as e:  # pragma: no cover
            logging.exception("Failed %s: %s", name, e)
            total_failed += 1

    if batch:
        res = _flush_batch(search, batch)
        total_uploaded += res["uploaded"]
        total_failed += res["failed"]

    return {"blobs": total_blobs, "chunks": total_chunks, "uploaded": total_uploaded, "failed": total_failed}
